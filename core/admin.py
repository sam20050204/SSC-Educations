from django.contrib import admin
from .models import Student, Enquiry, Admission, Payment, Bill, BillItem

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ['name', 'email', 'mobile', 'course', 'admission_date', 'is_active']
    list_filter = ['is_active', 'course', 'admission_date']
    search_fields = ['name', 'email', 'mobile']
    readonly_fields = ['student_id', 'date_registered', 'password']
    
    fieldsets = (
        ('Personal Information', {
            'fields': ('name', 'email', 'mobile')
        }),
        ('Course Information', {
            'fields': ('course', 'admission_date')
        }),
        ('Account Information', {
            'fields': ('student_id', 'is_active', 'date_registered')
        }),
        ('Security', {
            'fields': ('password',),
            'classes': ('collapse',)
        })
    )
    
    def get_readonly_fields(self, request, obj=None):
        if obj:  # Editing an existing object
            return self.readonly_fields + ['email', 'mobile']
        return self.readonly_fields

@admin.register(Enquiry)
class EnquiryAdmin(admin.ModelAdmin):
    list_display = ['enquiry_no', 'student_name', 'mobile_no', 'course', 'enquiry_date', 'created_at']
    list_filter = ['course', 'enquiry_date', 'created_at']
    search_fields = ['enquiry_no', 'student_name', 'mobile_no']
    readonly_fields = ['enquiry_no', 'created_at', 'updated_at']
    ordering = ['-created_at']
    
    fieldsets = (
        ('Enquiry Information', {
            'fields': ('enquiry_no', 'enquiry_date')
        }),
        ('Student Details', {
            'fields': ('student_name', 'mobile_no', 'course', 'address')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

@admin.register(Admission)
class AdmissionAdmin(admin.ModelAdmin):
    list_display = [
        'form_no', 
        'get_full_name', 
        'course_name',
        'batch',
        'mobile_own', 
        'admission_date',
        'installments',
        'created_by',
        'is_active'
    ]
    
    list_filter = [
        'is_active', 
        'course_name',
        'batch',
        'installments',
        'admission_date',
        'created_at'
    ]
    
    search_fields = [
        'form_no',
        'first_name',
        'middle_name', 
        'last_name',
        'mobile_own',
        'mobile_parents',
        'address'
    ]
    
    readonly_fields = [
        'form_no',
        'created_at',
        'updated_at',
        'created_by'
    ]
    
    ordering = ['-created_at']
    
    date_hierarchy = 'admission_date'
    
    fieldsets = (
        ('Form Information', {
            'fields': ('form_no', 'admission_date', 'batch', 'course_name')
        }),
        ('Student Details', {
            'fields': (
                'first_name',
                'middle_name',
                'last_name',
                'birth_date',
                'qualification'
            )
        }),
        ('Contact Information', {
            'fields': ('mobile_own', 'mobile_parents', 'address')
        }),
        ('Fee Information', {
            'fields': ('total_fees', 'paid_fees', 'installments')
        }),
        ('Photo', {
            'fields': ('photo',),
            'classes': ('collapse',)
        }),
        ('System Information', {
            'fields': ('created_by', 'is_active', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_readonly_fields(self, request, obj=None):
        if obj:  # Editing an existing object
            return self.readonly_fields + ['admission_date', 'mobile_own']
        return self.readonly_fields
    
    def get_full_name(self, obj):
        """Display full name in admin list"""
        return obj.get_full_name()
    get_full_name.short_description = 'Student Name'
    get_full_name.admin_order_field = 'first_name'
    
    # Custom actions
    actions = ['mark_inactive', 'mark_active', 'export_to_excel']
    
    def mark_inactive(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} admission(s) marked as inactive.')
    mark_inactive.short_description = 'Mark selected admissions as inactive'
    
    def mark_active(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} admission(s) marked as active.')
    mark_active.short_description = 'Mark selected admissions as active'
    
    def export_to_excel(self, request, queryset):
        """Export selected admissions to Excel"""
        import openpyxl
        from django.http import HttpResponse
        from datetime import datetime
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admissions"
        
        # Headers
        headers = [
            'Form No', 'Admission Date', 'Course', 'Batch',
            'First Name', 'Middle Name', 'Last Name',
            'Birth Date', 'Mobile (Own)', 'Mobile (Parents)',
            'Address', 'Qualification', 'Total Fees', 'Paid Fees',
            'Installments', 'Created By', 'Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        # Write data
        for row, admission in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=admission.form_no)
            ws.cell(row=row, column=2, value=admission.admission_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=admission.course_name)
            ws.cell(row=row, column=4, value=admission.batch)
            ws.cell(row=row, column=5, value=admission.first_name)
            ws.cell(row=row, column=6, value=admission.middle_name)
            ws.cell(row=row, column=7, value=admission.last_name)
            ws.cell(row=row, column=8, value=admission.birth_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=9, value=admission.mobile_own)
            ws.cell(row=row, column=10, value=admission.mobile_parents or 'N/A')
            ws.cell(row=row, column=11, value=admission.address)
            ws.cell(row=row, column=12, value=admission.qualification)
            ws.cell(row=row, column=13, value=float(admission.total_fees))
            ws.cell(row=row, column=14, value=float(admission.paid_fees))
            ws.cell(row=row, column=15, value=admission.get_installments_display())
            ws.cell(row=row, column=16, value=admission.created_by or 'N/A')
            ws.cell(row=row, column=17, value='Active' if admission.is_active else 'Inactive')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"admissions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook
        wb.save(response)
        
        self.message_user(request, f'Exported {queryset.count()} admission(s) to Excel.')
        return response
    
    export_to_excel.short_description = 'Export selected admissions to Excel'


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ['receipt_no', 'get_student_name', 'payment_date', 'amount_paid', 'payment_mode', 'created_at']
    list_filter = ['payment_mode', 'payment_date', 'created_at']
    search_fields = ['receipt_no', 'admission__first_name', 'admission__last_name', 'transaction_ref']
    readonly_fields = ['receipt_no', 'created_at', 'updated_at']
    ordering = ['-payment_date', '-created_at']
    date_hierarchy = 'payment_date'
    
    fieldsets = (
        ('Payment Information', {
            'fields': ('receipt_no', 'payment_date', 'admission')
        }),
        ('Amount Details', {
            'fields': ('amount_paid', 'payment_mode', 'transaction_ref')
        }),
        ('Additional Information', {
            'fields': ('remarks', 'created_by')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_student_name(self, obj):
        return obj.admission.get_full_name()
    get_student_name.short_description = 'Student Name'
    get_student_name.admin_order_field = 'admission__first_name'


@admin.register(Bill)
class BillAdmin(admin.ModelAdmin):
    list_display = ['receipt_no', 'bill_date', 'customer_name', 'customer_mobile', 'get_items_count', 'total_amount', 'created_at']
    list_filter = ['bill_date', 'created_at']
    search_fields = ['receipt_no', 'customer_name', 'customer_mobile']
    readonly_fields = ['receipt_no', 'created_at', 'updated_at']
    ordering = ['-bill_date', '-created_at']
    date_hierarchy = 'bill_date'
    
    fieldsets = (
        ('Bill Information', {
            'fields': ('receipt_no', 'bill_date')
        }),
        ('Customer Details', {
            'fields': ('customer_name', 'customer_mobile')
        }),
        ('Amount', {
            'fields': ('total_amount',)
        }),
        ('System Information', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_items_count(self, obj):
        return obj.items.count()
    get_items_count.short_description = 'Items Count'


@admin.register(BillItem)
class BillItemAdmin(admin.ModelAdmin):
    list_display = ['get_receipt_no', 'item_name', 'quantity', 'rate', 'amount', 'created_at']
    list_filter = ['created_at']
    search_fields = ['bill__receipt_no', 'item_name']
    readonly_fields = ['amount', 'created_at']
    ordering = ['-created_at']
    
    fieldsets = (
        ('Bill Reference', {
            'fields': ('bill',)
        }),
        ('Item Details', {
            'fields': ('item_name', 'quantity', 'rate', 'amount')
        }),
        ('Timestamp', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        })
    )
    
    def get_receipt_no(self, obj):
        return obj.bill.receipt_no
    get_receipt_no.short_description = 'Receipt No'
    get_receipt_no.admin_order_field = 'bill__receipt_no'