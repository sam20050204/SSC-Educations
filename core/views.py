# core/views.py - Complete file with all views

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models, transaction
from django.db.models import Q, Sum, Count
from .models import Student, Enquiry, Admission, Payment, Bill, BillItem
from datetime import datetime, timedelta
from decimal import Decimal
import json
import openpyxl
from collections import defaultdict


# ==================== HOME & AUTH VIEWS ====================

def home(request):
    """Home page view"""
    return render(request, 'home.html')


def register(request):
    """Student registration view"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')
        
        # Validation
        if not all([name, mobile, email, password1, password2]):
            return render(request, 'register.html', {
                'error': 'All fields are required',
                'form_data': request.POST
            })
        
        if password1 != password2:
            return render(request, 'register.html', {
                'error': 'Passwords do not match',
                'form_data': request.POST
            })
        
        if len(password1) < 6:
            return render(request, 'register.html', {
                'error': 'Password must be at least 6 characters long',
                'form_data': request.POST
            })
        
        if len(mobile) != 10 or not mobile.isdigit():
            return render(request, 'register.html', {
                'error': 'Mobile number must be 10 digits',
                'form_data': request.POST
            })
        
        # Check if email or mobile already exists
        if Student.objects.filter(email=email).exists():
            return render(request, 'register.html', {
                'error': 'Email already registered',
                'form_data': request.POST
            })
        
        if Student.objects.filter(mobile=mobile).exists():
            return render(request, 'register.html', {
                'error': 'Mobile number already registered',
                'form_data': request.POST
            })
        
        # Create student
        try:
            student = Student.objects.create(
                name=name,
                mobile=mobile,
                email=email,
                password=password1  # Will be hashed in model's save method
            )
            messages.success(request, 'Registration successful! Please login.')
            return redirect('login')
        except Exception as e:
            return render(request, 'register.html', {
                'error': f'Registration failed: {str(e)}',
                'form_data': request.POST
            })
    
    return render(request, 'register.html')


def login_view(request):
    """Login view"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        
        if not email or not password:
            return render(request, 'login.html', {
                'error': 'Email and password are required'
            })
        
        try:
            student = Student.objects.get(email=email, is_active=True)
            
            if check_password(password, student.password):
                # Set session
                request.session['student_id'] = str(student.student_id)
                request.session['student_name'] = student.name
                request.session['student_email'] = student.email
                
                messages.success(request, f'Welcome back, {student.name}!')
                return redirect('dashboard')
            else:
                return render(request, 'login.html', {
                    'error': 'Invalid email or password'
                })
        except Student.DoesNotExist:
            return render(request, 'login.html', {
                'error': 'Invalid email or password'
            })
    
    return render(request, 'login.html')


def logout_view(request):
    """Logout view"""
    request.session.flush()
    messages.success(request, 'Logged out successfully')
    return redirect('login')


def forgot_password(request):
    """Forgot password view"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        if not all([email, new_password, confirm_password]):
            return render(request, 'forgot_password.html', {
                'error': 'All fields are required'
            })
        
        if new_password != confirm_password:
            return render(request, 'forgot_password.html', {
                'error': 'Passwords do not match'
            })
        
        if len(new_password) < 6:
            return render(request, 'forgot_password.html', {
                'error': 'Password must be at least 6 characters long'
            })
        
        try:
            student = Student.objects.get(email=email)
            student.password = make_password(new_password)
            student.save()
            
            messages.success(request, 'Password reset successful! Please login with your new password.')
            return redirect('login')
        except Student.DoesNotExist:
            return render(request, 'forgot_password.html', {
                'error': 'Email not found'
            })
    
    return render(request, 'forgot_password.html')


# ==================== DASHBOARD ====================

def dashboard(request):
    """Dashboard view"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access dashboard')
        return redirect('login')
    
    # Get selected year
    selected_year = int(request.GET.get('year', datetime.now().year))
    
    # Get all admissions for the selected year
    admissions = Admission.objects.filter(
        admission_date__year=selected_year,
        is_active=True
    )
    
    # Total statistics
    total_students = admissions.count()
    mscit_count = admissions.filter(course_name='MS-CIT').count()
    klic_count = admissions.exclude(course_name='MS-CIT').count()
    
    # Monthly data for bar chart
    monthly_data = [0] * 12
    for admission in admissions:
        month = admission.admission_date.month - 1
        monthly_data[month] += 1
    
    # Course distribution for pie chart
    course_counts = admissions.values('course_name').annotate(count=Count('id'))
    pie_data = [{'label': item['course_name'], 'value': item['count']} for item in course_counts]
    
    # Available years
    years_range = range(datetime.now().year, datetime.now().year - 5, -1)
    
    context = {
        'student_name': request.session.get('student_name'),
        'total_students': total_students,
        'mscit_count': mscit_count,
        'klic_count': klic_count,
        'monthly_data': json.dumps(monthly_data),
        'months': json.dumps(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']),
        'pie_data': json.dumps(pie_data),
        'selected_year': selected_year,
        'available_years': years_range
    }
    
    return render(request, 'dashboard.html', context)


# ==================== ENQUIRY VIEWS ====================

def new_enquiry(request):
    """New enquiry form view"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    if request.method == 'POST':
        student_name = request.POST.get('student_name', '').strip()
        mobile_no = request.POST.get('mobile_no', '').strip()
        course = request.POST.get('course', '')
        address = request.POST.get('address', '').strip()
        
        # Validation
        if not all([student_name, mobile_no, course, address]):
            return render(request, 'new_enquiry.html', {
                'error': 'All fields are required',
                'form_data': request.POST,
                'student_name': request.session.get('student_name'),
                'current_date': datetime.now().strftime('%Y-%m-%d')
            })
        
        if len(mobile_no) != 10 or not mobile_no.isdigit():
            return render(request, 'new_enquiry.html', {
                'error': 'Mobile number must be 10 digits',
                'form_data': request.POST,
                'student_name': request.session.get('student_name'),
                'current_date': datetime.now().strftime('%Y-%m-%d')
            })
        
        # Create enquiry
        try:
            enquiry = Enquiry.objects.create(
                student_name=student_name,
                mobile_no=mobile_no,
                course=course,
                address=address
            )
            messages.success(request, f'Enquiry created successfully! Enquiry No: {enquiry.enquiry_no}')
            return redirect('enquiry_data')
        except Exception as e:
            return render(request, 'new_enquiry.html', {
                'error': f'Failed to create enquiry: {str(e)}',
                'form_data': request.POST,
                'student_name': request.session.get('student_name'),
                'current_date': datetime.now().strftime('%Y-%m-%d')
            })
    
    return render(request, 'new_enquiry.html', {
        'student_name': request.session.get('student_name'),
        'current_date': datetime.now().strftime('%Y-%m-%d')
    })


def enquiry_data(request):
    """View all enquiries"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    enquiries = Enquiry.objects.all().order_by('-created_at')
    
    return render(request, 'enquiry_data.html', {
        'student_name': request.session.get('student_name'),
        'enquiries': enquiries,
        'total_enquiries': enquiries.count()
    })


def export_enquiries(request):
    """Export enquiries to Excel"""
    if 'student_id' not in request.session:
        return redirect('login')
    
    enquiries = Enquiry.objects.all().order_by('-created_at')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    
    # Headers
    headers = ['Enquiry No', 'Date', 'Student Name', 'Mobile No', 'Course', 'Address']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
    
    # Data
    for row, enquiry in enumerate(enquiries, 2):
        ws.cell(row=row, column=1, value=enquiry.enquiry_no)
        ws.cell(row=row, column=2, value=enquiry.enquiry_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=enquiry.student_name)
        ws.cell(row=row, column=4, value=enquiry.mobile_no)
        ws.cell(row=row, column=5, value=enquiry.get_course_display())
        ws.cell(row=row, column=6, value=enquiry.address)
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="enquiries_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


# ==================== ADMISSION VIEWS ====================

def new_admission(request):
    """New admission form view"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    if request.method == 'POST':
        try:
            # Get form data
            admission_date = request.POST.get('admission_date')
            batch = request.POST.get('batch')
            course_name = request.POST.get('course_name')
            first_name = request.POST.get('first_name', '').strip()
            middle_name = request.POST.get('middle_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            birth_date = request.POST.get('birth_date')
            mobile_own = request.POST.get('mobile_own', '').strip()
            mobile_parents = request.POST.get('mobile_parents', '').strip()
            address = request.POST.get('address', '').strip()
            qualification = request.POST.get('qualification', '').strip()
            installment = request.POST.get('installment')
            photo = request.FILES.get('photo')
            
            # Validation
            if not all([admission_date, batch, course_name, first_name, middle_name, last_name,
                       birth_date, mobile_own, address, qualification, installment]):
                messages.error(request, 'All required fields must be filled')
                return render(request, 'new_admission.html', {
                    'error': 'All required fields must be filled',
                    'form_data': request.POST,
                    'student_name': request.session.get('student_name')
                })
            
            if len(mobile_own) != 10 or not mobile_own.isdigit():
                messages.error(request, 'Mobile number must be 10 digits')
                return render(request, 'new_admission.html', {
                    'error': 'Mobile number must be 10 digits',
                    'form_data': request.POST,
                    'student_name': request.session.get('student_name')
                })
            
            # Create admission
            admission = Admission.objects.create(
                admission_date=admission_date,
                batch=batch,
                course_name=course_name,
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                birth_date=birth_date,
                mobile_own=mobile_own,
                mobile_parents=mobile_parents if mobile_parents else None,
                address=address,
                qualification=qualification,
                installments=installment,
                photo=photo,
                created_by=request.session.get('student_name')
            )
            
            messages.success(request, f'Admission created successfully! Form No: {admission.form_no}')
            return redirect('admitted_students')
            
        except Exception as e:
            messages.error(request, f'Error creating admission: {str(e)}')
            return render(request, 'new_admission.html', {
                'error': str(e),
                'form_data': request.POST,
                'student_name': request.session.get('student_name')
            })
    
    return render(request, 'new_admission.html', {
        'student_name': request.session.get('student_name')
    })


def admitted_students(request):
    """View admitted students"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    return render(request, 'admitted_students.html', {
        'student_name': request.session.get('student_name')
    })


def students_details(request):
    """View all registered students"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    students = Student.objects.all().order_by('-date_registered')
    
    return render(request, 'students_details.html', {
        'student_name': request.session.get('student_name'),
        'students': students
    })


# ==================== API ENDPOINTS ====================

@csrf_exempt
def get_admitted_students(request):
    """API: Get admitted students by course and batch"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    try:
        course = request.GET.get('course', '')
        batch = request.GET.get('batch', '')
        
        if not course or not batch:
            return JsonResponse({
                'success': False,
                'error': 'Course and batch are required'
            })
        
        students = Admission.objects.filter(
            course_name=course,
            batch=batch,
            is_active=True
        ).order_by('-created_at')
        
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'formNo': student.form_no,
                'admissionDate': student.admission_date.strftime('%Y-%m-%d'),
                'course': student.course_name,
                'batch': student.batch,
                'firstName': student.first_name,
                'middleName': student.middle_name,
                'lastName': student.last_name,
                'birthDate': student.birth_date.strftime('%Y-%m-%d'),
                'mobileOwn': student.mobile_own,
                'mobileParents': student.mobile_parents or '',
                'address': student.address,
                'qualification': student.qualification,
                'installments': student.installments,
                'totalFees': float(student.total_fees),
                'paidFees': float(student.paid_fees),
                'remainingFees': float(student.get_remaining_fees()),
                'photo': student.photo.url if student.photo else None
            })
        
        return JsonResponse({
            'success': True,
            'students': students_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
def update_student(request):
    """API: Update student data"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('id')
            
            student = Admission.objects.get(id=student_id)
            
            # Update fields
            student.first_name = data.get('firstName')
            student.middle_name = data.get('middleName')
            student.last_name = data.get('lastName')
            student.birth_date = data.get('birthDate')
            student.mobile_own = data.get('mobileOwn')
            student.mobile_parents = data.get('mobileParents') or None
            student.address = data.get('address')
            student.qualification = data.get('qualification')
            student.total_fees = Decimal(str(data.get('totalFees')))
            student.paid_fees = Decimal(str(data.get('paidFees')))
            
            student.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Student updated successfully'
            })
            
        except Admission.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Student not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_exempt
def delete_student_admission(request):
    """API: Delete student admission"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            admission_id = data.get('admission_id')
            
            admission = Admission.objects.get(id=admission_id)
            student_name = admission.get_full_name()
            
            # Delete admission (will also delete related payments due to CASCADE)
            admission.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully deleted admission for {student_name}'
            })
            
        except Admission.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Admission not found'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


# ==================== PAYMENT VIEWS ====================

def fees_payment(request):
    """Fees payment view"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    if request.method == 'POST':
        try:
            admission_id = request.POST.get('admission_id')
            amount_paid = request.POST.get('amount_paid')
            payment_mode = request.POST.get('payment_mode')
            transaction_ref = request.POST.get('transaction_ref', '').strip()
            remarks = request.POST.get('remarks', '').strip()
            
            # Validation
            if not all([admission_id, amount_paid, payment_mode]):
                messages.error(request, 'All required fields must be filled')
                return redirect('fees_payment')
            
            admission = Admission.objects.get(id=admission_id)
            amount = Decimal(amount_paid)
            
            if amount <= 0:
                messages.error(request, 'Amount must be greater than zero')
                return redirect('fees_payment')
            
            if amount > admission.get_remaining_fees():
                messages.error(request, 'Amount cannot exceed remaining fees')
                return redirect('fees_payment')
            
            # Create payment
            payment = Payment.objects.create(
                payment_date=datetime.now().date(),
                admission=admission,
                amount_paid=amount,
                payment_mode=payment_mode,
                transaction_ref=transaction_ref if transaction_ref else None,
                remarks=remarks if remarks else None,
                created_by=request.session.get('student_name')
            )
            
            messages.success(request, f'Payment recorded successfully! Receipt No: {payment.receipt_no}')
            return redirect(f'/fees-payment/?receipt={payment.receipt_no}')
            
        except Admission.DoesNotExist:
            messages.error(request, 'Student not found')
            return redirect('fees_payment')
        except Exception as e:
            messages.error(request, f'Error processing payment: {str(e)}')
            return redirect('fees_payment')
    
    return render(request, 'fees_payment.html', {
        'student_name': request.session.get('student_name'),
        'today': datetime.now().strftime('%Y-%m-%d')
    })


def payment_history(request):
    """Payment history view"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page')
        return redirect('login')
    
    return render(request, 'payment_history.html', {
        'student_name': request.session.get('student_name')
    })


@csrf_exempt
def search_student_for_payment(request):
    """API: Search student for payment"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            search_term = data.get('search_term', '').strip()
            
            if len(search_term) < 2:
                return JsonResponse({
                    'success': True,
                    'students': []
                })
            
            # Search by name or mobile
            students = Admission.objects.filter(
                Q(first_name__icontains=search_term) |
                Q(middle_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(mobile_own__icontains=search_term),
                is_active=True
            )[:10]
            
            students_data = []
            for student in students:
                remaining = student.get_remaining_fees()
                if remaining > 0:  # Only show students with pending fees
                    students_data.append({
                        'id': student.id,
                        'form_no': student.form_no,
                        'full_name': student.get_full_name(),
                        'mobile': student.mobile_own,
                        'course': student.course_name,
                        'batch': student.batch,
                        'total_fees': float(student.total_fees),
                        'paid_fees': float(student.paid_fees),
                        'remaining_fees': float(remaining)
                    })
            
            return JsonResponse({
                'success': True,
                'students': students_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@csrf_exempt
def get_payment_history(request):
    """API: Get payment history"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    try:
        course = request.GET.get('course', '')
        batch = request.GET.get('batch', '')
        student_name = request.GET.get('student_name', '').strip()
        
        payments = Payment.objects.select_related('admission').all()
        
        if course:
            payments = payments.filter(admission__course_name=course)
        if batch:
            payments = payments.filter(admission__batch=batch)
        if student_name:
            payments = payments.filter(
                Q(admission__first_name__icontains=student_name) |
                Q(admission__middle_name__icontains=student_name) |
                Q(admission__last_name__icontains=student_name)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment History"
        
        # Headers
        headers = [
            'Receipt No', 'Payment Date', 'Student Name', 'Form No',
            'Course', 'Batch', 'Amount Paid', 'Payment Mode',
            'Transaction Ref', 'Remarks', 'Created By'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
        
        # Data
        for row, payment in enumerate(payments, 2):
            ws.cell(row=row, column=1, value=payment.receipt_no)
            ws.cell(row=row, column=2, value=payment.payment_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=payment.admission.get_full_name())
            ws.cell(row=row, column=4, value=payment.admission.form_no)
            ws.cell(row=row, column=5, value=payment.admission.course_name)
            ws.cell(row=row, column=6, value=payment.admission.batch)
            ws.cell(row=row, column=7, value=float(payment.amount_paid))
            ws.cell(row=row, column=8, value=payment.payment_mode)
            ws.cell(row=row, column=9, value=payment.transaction_ref or 'N/A')
            ws.cell(row=row, column=10, value=payment.remarks or 'N/A')
            ws.cell(row=row, column=11, value=payment.created_by or 'N/A')
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payment_history_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting payment history: {str(e)}')
        return redirect('payment_history')


# ==================== BILL MANAGEMENT VIEWS ====================

def new_bill(request):
    """Create new bill page"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page.')
        return redirect('login')
    
    if request.method == 'POST':
        try:
            # Get form data
            bill_date = request.POST.get('bill_date', '').strip()
            customer_name = request.POST.get('customer_name', '').strip()
            customer_mobile = request.POST.get('customer_mobile', '').strip()
            items_json = request.POST.get('items', '[]')
            
            # Validation
            if not bill_date or not customer_name or not customer_mobile:
                return JsonResponse({
                    'success': False,
                    'error': 'All fields are required'
                })
            
            if len(customer_mobile) != 10 or not customer_mobile.isdigit():
                return JsonResponse({
                    'success': False,
                    'error': 'Mobile number must be 10 digits'
                })
            
            # Parse items
            items = json.loads(items_json)
            if not items or len(items) == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'At least one item is required'
                })
            
            # Calculate total
            total_amount = Decimal('0.00')
            for item in items:
                amount = Decimal(str(item['amount']))
                total_amount += amount
            
            # Create bill and items in a transaction
            with transaction.atomic():
                # Create bill
                bill = Bill.objects.create(
                    bill_date=bill_date,
                    customer_name=customer_name,
                    customer_mobile=customer_mobile,
                    total_amount=total_amount,
                    created_by=request.session.get('student_name', 'Admin')
                )
                
                # Create bill items
                for item in items:
                    BillItem.objects.create(
                        bill=bill,
                        item_name=item['item_name'],
                        quantity=Decimal(str(item['quantity'])),
                        rate=Decimal(str(item['rate'])),
                        amount=Decimal(str(item['amount']))
                    )
            
            return JsonResponse({
                'success': True,
                'bill_id': bill.id,
                'receipt_no': bill.receipt_no,
                'message': 'Bill created successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # GET request
    return render(request, 'new_bill.html', {
        'student_name': request.session.get('student_name')
    })


def bills_list(request):
    """View all bills page"""
    if 'student_id' not in request.session:
        messages.error(request, 'Please login to access this page.')
        return redirect('login')
    
    return render(request, 'bills.html', {
        'student_name': request.session.get('student_name')
    })


@csrf_exempt
def get_bills(request):
    """API endpoint to get filtered bills"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    try:
        date = request.GET.get('date', '')
        customer = request.GET.get('customer', '').strip().lower()
        
        if not date:
            return JsonResponse({
                'success': False,
                'error': 'Date is required'
            })
        
        # Filter bills by date
        bills = Bill.objects.filter(bill_date=date)
        
        # Filter by customer name if provided
        if customer:
            bills = bills.filter(
                models.Q(customer_name__icontains=customer)
            )
        
        # Prepare bills data
        bills_data = []
        for bill in bills:
            items_count = bill.items.count()
            bills_data.append({
                'id': bill.id,
                'receipt_no': bill.receipt_no,
                'bill_date': bill.bill_date.strftime('%Y-%m-%d'),
                'customer_name': bill.customer_name,
                'customer_mobile': bill.customer_mobile,
                'total_amount': float(bill.total_amount),
                'items_count': items_count
            })
        
        return JsonResponse({
            'success': True,
            'bills': bills_data,
            'count': len(bills_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def print_bill(request, bill_id):
    """Print bill page"""
    if 'student_id' not in request.session:
        return redirect('login')
    
    try:
        bill = get_object_or_404(Bill, id=bill_id)
        items = bill.items.all()
        
        # Convert amount to words
        amount_in_words = convert_amount_to_words(float(bill.total_amount))
        
        context = {
            'bill': bill,
            'items': items,
            'amount_in_words': amount_in_words
        }
        
        return render(request, 'print_bill.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading bill: {str(e)}')
        return redirect('bills_list')


def export_bills(request):
    """Export bills to Excel"""
    if 'student_id' not in request.session:
        return redirect('login')
    
    try:
        date = request.GET.get('date', '')
        customer = request.GET.get('customer', '').strip().lower()
        
        if not date:
            messages.error(request, 'Date is required for export')
            return redirect('bills_list')
        
        # Filter bills by date
        bills = Bill.objects.filter(bill_date=date)
        
        # Filter by customer name if provided
        if customer:
            bills = bills.filter(
                models.Q(customer_name__icontains=customer)
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: Bills Summary
        ws_summary = wb.active
        ws_summary.title = "Bills Summary"
        
        # Headers for summary
        summary_headers = [
            'Receipt No', 'Date', 'Customer Name', 
            'Mobile', 'Items Count', 'Total Amount'
        ]
        
        # Write summary headers
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
        
        # Write summary data
        row = 2
        for bill in bills:
            items_count = bill.items.count()
            ws_summary.cell(row=row, column=1, value=bill.receipt_no)
            ws_summary.cell(row=row, column=2, value=bill.bill_date.strftime('%d/%m/%Y'))
            ws_summary.cell(row=row, column=3, value=bill.customer_name)
            ws_summary.cell(row=row, column=4, value=bill.customer_mobile)
            ws_summary.cell(row=row, column=5, value=items_count)
            ws_summary.cell(row=row, column=6, value=float(bill.total_amount))
            row += 1
        
        # Auto-adjust column widths for summary
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Items Details
        ws_items = wb.create_sheet(title="Items Details")
        
        # Headers for items
        items_headers = [
            'Receipt No', 'Bill Date', 'Customer Name',
            'Item Name', 'Quantity', 'Rate', 'Amount'
        ]
        
        # Write items headers
        for col, header in enumerate(items_headers, 1):
            cell = ws_items.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="27AE60",
                end_color="27AE60",
                fill_type="solid"
            )
        
        # Write items data
        row = 2
        for bill in bills:
            for item in bill.items.all():
                ws_items.cell(row=row, column=1, value=bill.receipt_no)
                ws_items.cell(row=row, column=2, value=bill.bill_date.strftime('%d/%m/%Y'))
                ws_items.cell(row=row, column=3, value=bill.customer_name)
                ws_items.cell(row=row, column=4, value=item.item_name)
                ws_items.cell(row=row, column=5, value=float(item.quantity))
                ws_items.cell(row=row, column=6, value=float(item.rate))
                ws_items.cell(row=row, column=7, value=float(item.amount))
                row += 1
        
        # Auto-adjust column widths for items
        for column in ws_items.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_items.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        filename = f"bills_{date_obj.strftime('%d%m%Y')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting bills: {str(e)}')
        return redirect('bills_list')


# ==================== HELPER FUNCTIONS ====================

def convert_amount_to_words(amount):
    """Convert numeric amount to words (Indian numbering system)"""
    def num_to_words(n):
        under_20 = ['Zero', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 
                   'Eight', 'Nine', 'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 
                   'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
        
        if n < 20:
            return under_20[n]
        elif n < 100:
            return tens[n // 10] + ('' if n % 10 == 0 else ' ' + under_20[n % 10])
        elif n < 1000:
            return under_20[n // 100] + ' Hundred' + ('' if n % 100 == 0 else ' ' + num_to_words(n % 100))
        elif n < 100000:
            return num_to_words(n // 1000) + ' Thousand' + ('' if n % 1000 == 0 else ' ' + num_to_words(n % 1000))
        elif n < 10000000:
            return num_to_words(n // 100000) + ' Lakh' + ('' if n % 100000 == 0 else ' ' + num_to_words(n % 100000))
        else:
            return num_to_words(n // 10000000) + ' Crore' + ('' if n % 10000000 == 0 else ' ' + num_to_words(n % 10000000))
    
    try:
        rupees = int(amount)
        paise = int((amount - rupees) * 100)
        
        words = num_to_words(rupees) + ' Rupees'
        if paise > 0:
            words += ' and ' + num_to_words(paise) + ' Paise'
        
        return words + ' Only'
    except:
        return 'Amount conversion error'
 = payments.filter(
                Q(admission__first_name__icontains=student_name) |
                Q(admission__middle_name__icontains=student_name) |
                Q(admission__last_name__icontains=student_name)
            )
        
        payments_data = []
        for payment in payments:
            payments_data.append({
                'id': payment.id,
                'receipt_no': payment.receipt_no,
                'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
                'student_name': payment.admission.get_full_name(),
                'form_no': payment.admission.form_no,
                'course': payment.admission.course_name,
                'batch': payment.admission.batch,
                'amount_paid': float(payment.amount_paid),
                'payment_mode': payment.payment_mode,
                'transaction_ref': payment.transaction_ref or '',
                'remarks': payment.remarks or '',
                'created_by': payment.created_by or 'N/A'
            })
        
        return JsonResponse({
            'success': True,
            'payments': payments_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
def get_receipt_details(request):
    """API: Get receipt details"""
    if 'student_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    try:
        receipt_no = request.GET.get('receipt_no', '')
        
        if not receipt_no:
            return JsonResponse({
                'success': False,
                'error': 'Receipt number is required'
            })
        
        payment = Payment.objects.select_related('admission').get(receipt_no=receipt_no)
        
        receipt_data = {
            'receipt_no': payment.receipt_no,
            'payment_date': payment.payment_date.strftime('%d/%m/%Y'),
            'student_name': payment.admission.get_full_name(),
            'course': payment.admission.course_name,
            'batch': payment.admission.batch,
            'amount_paid': float(payment.amount_paid),
            'payment_mode': payment.payment_mode,
            'transaction_ref': payment.transaction_ref or '',
            'amount_in_words': payment.get_amount_in_words()
        }
        
        return JsonResponse({
            'success': True,
            'receipt': receipt_data
        })
        
    except Payment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Receipt not found'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


def export_payment_history(request):
    """Export payment history to Excel"""
    if 'student_id' not in request.session:
        return redirect('login')
    
    try:
        course = request.GET.get('course', '')
        batch = request.GET.get('batch', '')
        student_name = request.GET.get('student_name', '')
        
        payments = Payment.objects.select_related('admission').all()
        
        if course:
            payments = payments.filter(admission__course_name=course)
        if batch:
            payments = payments.filter(admission__batch=batch)
        if student_name:
            payments